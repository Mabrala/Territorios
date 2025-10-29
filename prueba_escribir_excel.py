from openpyxl import load_workbook
import os

ruta = os.path.join(os.path.dirname(__file__), 'Rec_Entreg.xlsx')
codigo = "O12"
nombre = "Raquel Villarreal"
fecha = "29/10/2025"

def registrar_entrega(ruta_excel, codigo, nombre, fecha):
    """
    Escribe una nueva fila en la hoja 'Entregados' del archivo Excel.
    
    Parámetros:
        codigo (str): Código del territorio (ej. 'O1', 'H12'...)
        nombre (str): Nombre de la persona
        fecha (str | datetime): Fecha de entrega (puede ser string 'dd/mm/aaaa' o datetime)
    """
    # Cargar el libro y la hoja
    wb = load_workbook(ruta_excel)
    hoja = wb["ENTREGADOS"]
    
    hoja_recibidos = wb["RECIBIDOS"]
    fila_borrar = None
    
    for fila in range(1, hoja_recibidos.max_row + 1):
        celda = hoja_recibidos.cell(row=fila, column=1).value
        if celda == codigo:
            fila_borrar = fila
            break
    
    if fila_borrar:
        hoja_recibidos.delete_rows(fila_borrar)
    else:
        print(f"{codigo} no encontrado en 'RECIBIDOS'.")
        
    # Buscar la primera fila vacía
    fila = hoja.max_row + 1
    while all(cell.value is None for cell in hoja[fila]):
        fila -= 1
        
    # Escribir los datos
    hoja.cell(row=fila, column=1, value=codigo)
    hoja.cell(row=fila, column=2, value=nombre)
    hoja.cell(row=fila, column=3, value=fecha)

    # Guardar los cambios
    wb.save(ruta_excel)
    wb.close()

registrar_entrega(ruta, codigo, nombre, fecha)